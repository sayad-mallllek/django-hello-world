from datetime import datetime, timedelta
from django import views
from django.shortcuts import render
from django.http import HttpResponse
from django.db.models import Sum
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from .models import ShippingProvider
from orders.models import OrderBasket


class ShippingProviderAnalyze(views.generic.ListView):
    """
    View class for shipping provider analysis
    """
    admin = {}
    
    def get(self, request):
        ctx = self.admin.each_context(request) if hasattr(self.admin, 'each_context') else {}
        
        # Get date parameters from request
        date_from_str = request.GET.get('date_from', '')
        date_to_str = request.GET.get('date_to', '')
        
        # Set default dates if not provided
        if not date_from_str:
            date_from = datetime.today() - timedelta(days=30)  # Default to last 30 days
        else:
            date_from = datetime.strptime(date_from_str, '%Y-%m-%d')
            
        if not date_to_str:
            date_to = datetime.today()
        else:
            date_to = datetime.strptime(date_to_str, '%Y-%m-%d')
            # Make the end date inclusive by setting it to end of day
            date_to = date_to.replace(hour=23, minute=59, second=59)
        
        # Get shipping provider stats with sum of item weights
        provider_stats = []
        
        shipping_providers = ShippingProvider.objects.all()
        for provider in shipping_providers:
            total_weight = OrderBasket.objects.filter(
                shipping_provider=provider,
                created_at__gte=date_from,
                created_at__lte=date_to
            ).aggregate(total_weight=Sum('items_weight'))['total_weight'] or 0
            
            order_count = OrderBasket.objects.filter(
                shipping_provider=provider,
                created_at__gte=date_from,
                created_at__lte=date_to
            ).count()
            
            provider_stats.append({
                'provider_name': provider.name,
                'total_weight': total_weight,
                'order_count': order_count,
                'provider_id': provider.id
            })
        
        # Add context variables
        ctx.update({
            'provider_stats': provider_stats,
            'date_from': date_from,
            'date_to': date_to,
            'is_filtered': bool(date_from_str or date_to_str)
        })
        
        return render(request, "shipping-provider-analyze.html", ctx)


def export_shipping_provider_analyze(request):
    """
    Export shipping provider analysis to Excel
    """
    # Get date parameters from request
    date_from_str = request.GET.get('date_from', '')
    date_to_str = request.GET.get('date_to', '')
    
    # Set default dates if not provided
    if not date_from_str:
        date_from = datetime.today() - timedelta(days=30)
    else:
        date_from = datetime.strptime(date_from_str, '%Y-%m-%d')
        
    if not date_to_str:
        date_to = datetime.today()
    else:
        date_to = datetime.strptime(date_to_str, '%Y-%m-%d')
        date_to = date_to.replace(hour=23, minute=59, second=59)
    
    # Get shipping provider stats
    shipping_providers = ShippingProvider.objects.all()
    provider_stats = []
    
    for provider in shipping_providers:
        total_weight = OrderBasket.objects.filter(
            shipping_provider=provider,
            created_at__gte=date_from,
            created_at__lte=date_to
        ).aggregate(total_weight=Sum('items_weight'))['total_weight'] or 0
        
        order_count = OrderBasket.objects.filter(
            shipping_provider=provider,
            created_at__gte=date_from,
            created_at__lte=date_to
        ).count()
        
        provider_stats.append({
            'provider_name': provider.name,
            'total_weight': total_weight,
            'order_count': order_count,
        })
    
    # Create a workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Shipping Provider Report"
    
    # Add header with date range
    ws.merge_cells('A1:C1')
    header_cell = ws['A1']
    header_cell.value = f"Shipping Provider Analysis ({date_from.strftime('%Y-%m-%d')} to {date_to.strftime('%Y-%m-%d')})"
    header_cell.font = Font(size=14, bold=True)
    header_cell.alignment = Alignment(horizontal='center')
    
    # Create column headers
    headers = ['Provider Name', 'Total Weight (kg)', 'Order Count']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}3'] = header
        ws[f'{col_letter}3'].font = Font(bold=True)
        ws[f'{col_letter}3'].fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Add data rows
    row_num = 4
    for stat in provider_stats:
        ws[f'A{row_num}'] = stat['provider_name']
        ws[f'B{row_num}'] = stat['total_weight']
        ws[f'C{row_num}'] = stat['order_count']
        row_num += 1
    
    # Add summary row
    ws[f'A{row_num+1}'] = f"Total Providers: {len(provider_stats)}"
    ws[f'A{row_num+1}'].font = Font(bold=True)
    
    ws[f'B{row_num+1}'] = sum(stat['total_weight'] for stat in provider_stats)
    ws[f'B{row_num+1}'].font = Font(bold=True)
    
    ws[f'C{row_num+1}'] = sum(stat['order_count'] for stat in provider_stats)
    ws[f'C{row_num+1}'].font = Font(bold=True)
    
    # Auto-size columns
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"shipping_provider_report_{date_from.strftime('%Y%m%d')}_to_{date_to.strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Save the workbook to the response
    wb.save(response)
    return response
