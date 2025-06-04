from datetime import datetime
from django import views
from django.shortcuts import render
from django.http import FileResponse, HttpResponse
from django.db.models import Prefetch
import io

from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

import arabic_reshaper
from bidi.algorithm import get_display

from .models import Order, OrderBasket


def process_arabic_text(text):
    """
    Process Arabic text for proper display in PDF by reshaping and applying bidirectional algorithm
    """
    if not text or not isinstance(text, str):
        return text
    
    # Reshape Arabic text to connect letters properly
    reshaped_text = arabic_reshaper.reshape(text)
    
    # Apply bidirectional algorithm for proper text direction
    bidi_text = get_display(reshaped_text)
    
    return bidi_text


def register_fonts():
    """
    Register fonts that support Arabic text
    """
    try:
        # Register DejaVu Sans which supports Arabic
        pdfmetrics.registerFont(TTFont('DejaVuSans', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
        pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'))
        return True
    except Exception as e:
        print(f"Warning: Could not register Arabic font: {e}")
        return False


def print_order_baskets_pdf(request):
    """
    Generate a PDF report for selected order baskets
    """
    if request.method != "GET":
        return HttpResponse("Method not allowed", status=405)
    
    # Register Arabic-supporting fonts
    arabic_font_available = register_fonts()
    
    # Get the basket IDs from the URL parameter
    basket_ids_str = request.GET.get('ids', '')
    if not basket_ids_str:
        return HttpResponse("No basket IDs provided", status=400)
    
    try:
        basket_ids = [int(id.strip()) for id in basket_ids_str.split(',') if id.strip()]
    except ValueError:
        return HttpResponse("Invalid basket IDs", status=400)
    
    if not basket_ids:
        return HttpResponse("No valid basket IDs provided", status=400)
    
    # Get the order baskets
    baskets = OrderBasket.objects.filter(id__in=basket_ids).prefetch_related(
        Prefetch('order_set', to_attr='orders')
    ).order_by('id')
    
    if not baskets.exists():
        return HttpResponse("No baskets found", status=404)
    
    # Create a file-like buffer to receive PDF data
    buffer = io.BytesIO()
    
    # Create the PDF document
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    story = []
    
    # Get sample style sheet
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30,
        alignment=TA_CENTER,
        textColor=colors.darkblue
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        spaceAfter=12,
        spaceBefore=20,
        textColor=colors.darkblue
    )
    
    # Title
    story.append(Paragraph("Order Baskets Summary Report", title_style))
    story.append(Spacer(1, 20))
    
    # Summary statistics
    total_baskets = baskets.count()
    total_orders = sum([len(basket.orders) for basket in baskets])
    total_amount = sum([basket.total_price for basket in baskets])
    
    story.append(Paragraph("Summary Statistics", heading_style))
    
    summary_data = [
        ['Total Baskets:', str(total_baskets)],
        ['Total Orders:', str(total_orders)],
        ['Total Amount:', f'${total_amount:.2f}'],
        ['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
    ]
    
    summary_table = Table(summary_data, colWidths=[2*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'DejaVuSans-Bold' if arabic_font_available else 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.lightgrey),
    ]))
    
    story.append(summary_table)
    story.append(Spacer(1, 30))
    
    # Detailed basket information
    story.append(Paragraph("Basket Details", heading_style))
    
    for basket in baskets:
        # Basket header
        basket_title = f"Basket #{basket.id}"
        story.append(Paragraph(basket_title, styles['Heading3']))
        
        # Basket info
        basket_info = [
            ['Created:', basket.created_at.strftime('%Y-%m-%d %H:%M')],
            ['Orders Count:', str(len(basket.orders))],
            ['Total Price:', f'${basket.total_price:.2f}'],
        ]
        
        basket_info_table = Table(basket_info, colWidths=[1.5*inch, 2*inch])
        basket_info_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'DejaVuSans-Bold' if arabic_font_available else 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        
        story.append(basket_info_table)
        story.append(Spacer(1, 10))
        
        # Orders in this basket
        if len(basket.orders) > 0:
            story.append(Paragraph("Orders:", styles['Heading4']))
            
            order_data = [['Order ID', 'Items Link', 'Quantity', 'Price', 'Status']]
            
            for order in basket.orders:
                order_data.append([
                    str(order.id),
                    order.items_link if order.items_link else 'N/A',
                    str(order.number_of_items),
                    f'${order.total_price:.2f}',
                    order.status
                ])
            
            order_table = Table(order_data, colWidths=[0.8*inch, 2.5*inch, 0.8*inch, 0.8*inch, 1*inch])
            order_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'DejaVuSans-Bold' if arabic_font_available else 'Helvetica-Bold'),
                ('FONTNAME', (0, 1), (-1, -1), 'DejaVuSans' if arabic_font_available else 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.beige, colors.lightgrey]),
            ]))
            
            story.append(order_table)
        
        story.append(Spacer(1, 20))
    
    # Build the PDF
    doc.build(story)
    
    # FileResponse sets the Content-Disposition header so that browsers
    # present the option to save the file.
    buffer.seek(0)
    filename = f"order_baskets_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return FileResponse(buffer, as_attachment=True, filename=filename)


def print_orders_pdf(request):
    """
    Generate a PDF report for selected orders
    """
    if request.method != "GET":
        return HttpResponse("Method not allowed", status=405)
    
    # Register Arabic-supporting fonts
    arabic_font_available = register_fonts()
    
    # Get the order IDs from the URL parameter
    order_ids_str = request.GET.get('ids', '')
    if not order_ids_str:
        return HttpResponse("No order IDs provided", status=400)
    
    try:
        order_ids = [int(id.strip()) for id in order_ids_str.split(',') if id.strip()]
    except ValueError:
        return HttpResponse("Invalid order IDs", status=400)
    
    if not order_ids:
        return HttpResponse("No valid order IDs provided", status=400)
    
    # Get the orders with related data
    orders = Order.objects.filter(id__in=order_ids).select_related(
        'customer', 'order_basket', 'delivery_provider'
    ).order_by('id')
    
    if not orders.exists():
        return HttpResponse("No orders found", status=404)
    
    # Create a file-like buffer to receive PDF data
    buffer = io.BytesIO()
    
    # Create the PDF document
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    story = []
    
    # Get sample style sheet
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30,
        alignment=TA_CENTER,
        textColor=colors.darkblue
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        spaceAfter=12,
        spaceBefore=20,
        textColor=colors.darkblue
    )
    
    # Title
    story.append(Paragraph("Orders Summary Report", title_style))
    story.append(Spacer(1, 20))
    
    # Summary statistics
    total_orders = orders.count()
    total_amount = sum([order.total_price for order in orders])
    total_items = sum([order.number_of_items for order in orders])
    total_delivery_charges = sum([order.delivery_charge or 0 for order in orders])
    
    story.append(Paragraph("Summary Statistics", heading_style))
    
    summary_data = [
        ['Total Orders:', str(total_orders)],
        ['Total Items:', str(total_items)],
        ['Total Amount:', f'${total_amount:.2f}'],
        ['Total Delivery Charges:', f'${total_delivery_charges:.2f}'],
        ['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
    ]
    
    summary_table = Table(summary_data, colWidths=[2*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'DejaVuSans-Bold' if arabic_font_available else 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.lightgrey),
    ]))
    
    story.append(summary_table)
    story.append(Spacer(1, 30))
    
    # Detailed order information
    story.append(Paragraph("Order Details", heading_style))
    
    # Create table with order data
    order_data = [['Order ID', 'Customer', 'Bill ID', 'Status', 'Items', 'Total Price', 'Delivery Charge', 'Created At']]
    
    for order in orders:
        customer_name = order.customer.full_name if order.customer else 'N/A'
        processed_customer_name = process_arabic_text(customer_name) if customer_name != 'N/A' else 'N/A'
        
        order_data.append([
            str(order.id),
            processed_customer_name,
            order.bill_id if order.bill_id else 'N/A',
            order.get_status_display(),
            str(order.number_of_items),
            f'${order.total_price:.2f}',
            f'${order.delivery_charge:.2f}' if order.delivery_charge else '$0.00',
            order.created_at.strftime('%Y-%m-%d %H:%M')
        ])
    
    order_table = Table(order_data, colWidths=[0.7*inch, 1.5*inch, 0.8*inch, 0.8*inch, 0.6*inch, 0.8*inch, 0.8*inch, 1.1*inch])
    order_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'DejaVuSans-Bold' if arabic_font_available else 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'DejaVuSans' if arabic_font_available else 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.beige, colors.lightgrey]),
    ]))
    
    story.append(order_table)
    story.append(Spacer(1, 30))
    
    # Order details breakdown
    story.append(Paragraph("Individual Order Information", heading_style))
    
    for order in orders:
        # Order header
        order_title = f"Order #{order.id}"
        story.append(Paragraph(order_title, styles['Heading3']))
        
        # Order info
        customer_name = order.customer.full_name if order.customer else '-'
        processed_customer_name = process_arabic_text(customer_name) if customer_name != '-' else '-'
        
        customer_address = order.customer.address if order.customer else '-'
        processed_address = process_arabic_text(customer_address) if customer_address != '-' else '-'

        print("processed_address:", processed_address)
        
        order_info = [
            ['Shippier:', 'Finders - Shop And Ship'],
            ['Customer:', processed_customer_name],
            ['Note:', order.notes if order.notes else '-'],
            ['Tel:', order.customer.phone_number if order.customer else '-'],
            ['Address:', processed_address],
            ['Price:', f'${((order.total_price or 0) + (order.delivery_charge or 0)):.2f}'],
        ]
        
        if order.notes:
            order_info.append(['Notes:', order.notes[:100] + '...' if len(order.notes) > 100 else order.notes])
        
        order_info_table = Table(order_info, colWidths=[2*inch, 3*inch])
        order_info_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'DejaVuSans-Bold' if arabic_font_available else 'Helvetica-Bold'),
            ('FONTNAME', (0, 0), (1, -1), 'DejaVuSans' if arabic_font_available else 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ]))
        
        story.append(order_info_table)
        story.append(Spacer(1, 20))
    
    # Build the PDF
    doc.build(story)
    
    # FileResponse sets the Content-Disposition header so that browsers
    # present the option to save the file.
    buffer.seek(0)
    filename = f"orders_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return FileResponse(buffer, as_attachment=True, filename=filename)


def export_range_summary(request):
    """
    Export orders within a date range to Excel
    """
    from datetime import datetime, timedelta
    from django.http import HttpResponse
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    
    # Get date parameters from request
    date_from_str = request.GET.get('date_from', '')
    date_to_str = request.GET.get('date_to', '')
    
    # Set default dates if not provided
    if not date_from_str:
        date_from = datetime.today() - timedelta(days=30)
    else:
        date_from = datetime.strptime(date_from_str, '%Y-%m-%d')
        
    if not date_to_str:
        date_to = datetime.today()
    else:
        date_to = datetime.strptime(date_to_str, '%Y-%m-%d')
        date_to = date_to.replace(hour=23, minute=59, second=59)
    
    # Filter orders by date range - no pagination for export
    orders = Order.objects.filter(
        created_at__gte=date_from,
        created_at__lte=date_to
    ).order_by('-created_at')
    
    # Create a workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders Report"
    
    # Add header with date range
    ws.merge_cells('A1:F1')
    header_cell = ws['A1']
    header_cell.value = f"Orders Report ({date_from.strftime('%Y-%m-%d')} to {date_to.strftime('%Y-%m-%d')})"
    header_cell.font = Font(size=14, bold=True)
    header_cell.alignment = Alignment(horizontal='center')
    
    # Create column headers
    headers = ['ID', 'Customer', 'Status', 'Total Price ($)', 'Created At', 'Notes']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}3'] = header
        ws[f'{col_letter}3'].font = Font(bold=True)
        ws[f'{col_letter}3'].fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Add data rows
    row_num = 4
    for order in orders:
        ws[f'A{row_num}'] = order.id
        ws[f'B{row_num}'] = order.customer.full_name if order.customer else 'N/A'
        ws[f'C{row_num}'] = order.get_status_display()
        ws[f'D{row_num}'] = order.total_price
        ws[f'E{row_num}'] = order.created_at.strftime('%Y-%m-%d %H:%M')
        ws[f'F{row_num}'] = order.notes if hasattr(order, 'notes') else ''
        row_num += 1
    
    # Add summary row
    ws[f'A{row_num+1}'] = f"Total Orders: {orders.count()}"
    ws[f'A{row_num+1}'].font = Font(bold=True)
    
    ws[f'C{row_num+1}'] = "Total Value:"
    ws[f'C{row_num+1}'].font = Font(bold=True)
    
    ws[f'D{row_num+1}'] = sum(order.total_price for order in orders)
    ws[f'D{row_num+1}'].font = Font(bold=True)
    
    # Auto-size columns
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"orders_report_{date_from.strftime('%Y%m%d')}_to_{date_to.strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Save the workbook to the response
    wb.save(response)
    return response


# Range Summary view for order analysis
class RangeSummary(views.generic.ListView):
    """
    View class for order range summary analysis
    Can be used by admin interface
    """
    admin = {}
    
    def get(self, request):
        from datetime import datetime, timedelta
        from django.core.paginator import Paginator
        
        ctx = self.admin.each_context(request) if hasattr(self.admin, 'each_context') else {}
        
        # Get date parameters from request
        date_from_str = request.GET.get('date_from', '')
        date_to_str = request.GET.get('date_to', '')
        page_number = request.GET.get('page', 1)
        
        # Set default dates if not provided
        if not date_from_str:
            date_from = datetime.today() - timedelta(days=30)  # Default to last 30 days
        else:
            date_from = datetime.strptime(date_from_str, '%Y-%m-%d')
            
        if not date_to_str:
            date_to = datetime.today()
        else:
            date_to = datetime.strptime(date_to_str, '%Y-%m-%d')
            # Make the end date inclusive by setting it to end of day
            date_to = date_to.replace(hour=23, minute=59, second=59)
        
        # Filter orders by date range
        orders_query = Order.objects.filter(
            created_at__gte=date_from,
            created_at__lte=date_to
        ).order_by('-created_at')
        
        # Paginate the results
        paginator = Paginator(orders_query, 25)  # Show 25 orders per page
        orders_page = paginator.get_page(page_number)
        
        # Add context variables
        ctx.update({
            'orders': orders_page,
            'date_from': date_from,
            'date_to': date_to,
            'is_filtered': bool(date_from_str or date_to_str)
        })
        
        return render(request, "range-summary.html", ctx)
