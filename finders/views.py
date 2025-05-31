from datetime import datetime, timedelta
from django.shortcuts import render
from django import views
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.http import FileResponse, HttpResponse
from django.core.paginator import Paginator
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from django.template.loader import render_to_string
from django.http import JsonResponse
from orders.models import Order, OrderBasket
from expenses.models import Capital
from providers.models import ShippingProvider
from django.db.models import Sum, Count, Prefetch
import io


def generate_pdf(request):
    if request.method == "POST":

        # Create a file-like buffer to receive PDF data.
        buffer = io.BytesIO()

        # Create the PDF object, using the buffer as its "file."
        p = canvas.Canvas(buffer)

        # Draw things on the PDF. Here's where the PDF generation happens.
        # See the ReportLab documentation for the full list of functionality.
        p.drawString(100, 100, "Hello world.")

        # Close the PDF object cleanly, and we're done.
        p.showPage()
        p.save()

        # FileResponse sets the Content-Disposition header so that browsers
        # present the option to save the file.
        buffer.seek(0)
        return FileResponse(buffer, as_attachment=True, filename="hello.pdf")

    return HttpResponse("Not the correct method")


def print_order_baskets_pdf(request):
    """
    Generate a PDF report for selected order baskets
    """
    if request.method != "GET":
        return HttpResponse("Method not allowed", status=405)
    
    # Get the basket IDs from the URL parameter
    basket_ids_str = request.GET.get('ids', '')
    if not basket_ids_str:
        return HttpResponse("No basket IDs provided", status=400)
    
    try:
        basket_ids = [int(id.strip()) for id in basket_ids_str.split(',') if id.strip()]
    except ValueError:
        return HttpResponse("Invalid basket IDs", status=400)
    
    if not basket_ids:
        return HttpResponse("No valid basket IDs provided", status=400)
    
    # Get the order baskets
    baskets = OrderBasket.objects.filter(id__in=basket_ids).prefetch_related(Prefetch('order_set', to_attr='orders')).order_by('id')
    
    if not baskets.exists():
        return HttpResponse("No baskets found", status=404)
    
    # Create a file-like buffer to receive PDF data
    buffer = io.BytesIO()
    
    # Create the PDF document
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    story = []
    
    # Get sample style sheet
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30,
        alignment=TA_CENTER,
        textColor=colors.darkblue
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        spaceAfter=12,
        spaceBefore=20,
        textColor=colors.darkblue
    )
    
    # Title
    story.append(Paragraph("Order Baskets Summary Report", title_style))
    story.append(Spacer(1, 20))
    
    # Summary statistics
    total_baskets = baskets.count()
    total_orders = sum([len(basket.orders) for basket in baskets])
    total_amount = sum([basket.total_price for basket in baskets])
    
    story.append(Paragraph("Summary Statistics", heading_style))
    
    summary_data = [
        ['Total Baskets:', str(total_baskets)],
        ['Total Orders:', str(total_orders)],
        ['Total Amount:', f'${total_amount:.2f}'],
        ['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
    ]
    
    summary_table = Table(summary_data, colWidths=[2*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.lightgrey),
    ]))
    
    story.append(summary_table)
    story.append(Spacer(1, 30))
    
    # Detailed basket information
    story.append(Paragraph("Basket Details", heading_style))
    
    for basket in baskets:
        # Basket header
        basket_title = f"Basket #{basket.id}"
        story.append(Paragraph(basket_title, styles['Heading3']))
        
        # Basket info
        basket_info = [
            ['Created:', basket.created_at.strftime('%Y-%m-%d %H:%M')],
            ['Orders Count:', str(len(basket.orders))],
            ['Total Price:', f'${basket.total_price:.2f}'],
        ]
        
        basket_info_table = Table(basket_info, colWidths=[1.5*inch, 2*inch])
        basket_info_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        
        story.append(basket_info_table)
        story.append(Spacer(1, 10))
        
        # Orders in this basket
        if len(basket.orders) > 0:
            story.append(Paragraph("Orders:", styles['Heading4']))
            
            order_data = [['Order ID', 'Items Link', 'Quantity', 'Price', 'Status']]
            
            for order in basket.orders:
                print("Order", order)
                order_data.append([
                    str(order.id),
                    order.items_link if order.items_link else 'N/A',
                    str(order.number_of_items),
                    f'${order.total_price:.2f}',
                    order.status
                ])
            
            order_table = Table(order_data, colWidths=[0.8*inch, 2.5*inch, 0.8*inch, 0.8*inch, 1*inch])
            order_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.beige, colors.lightgrey]),
            ]))
            
            story.append(order_table)
        
        story.append(Spacer(1, 20))
    
    # Build the PDF
    doc.build(story)
    
    # FileResponse sets the Content-Disposition header so that browsers
    # present the option to save the file.
    buffer.seek(0)
    filename = f"order_baskets_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return FileResponse(buffer, as_attachment=True, filename=filename)
    

class Overview(views.generic.ListView):
    admin = {}

    def get(self, request):
        ctx = self.admin.each_context(request)
        
        ctx["capital"] = Capital.load()
        
        ctx["total_money_received_from_orders"] = (
            f"{Order.objects.get_all_received_money_from_orders()}$"
        )
        ctx["total_missing_money"] = (
            f"{Order.objects.get_missing_money_from_all_providers()}$"
        )
        ctx["email"] = "Email"
        return render(request, "overview.html", ctx)
class RangeSummary(views.generic.ListView):
    admin = {}
    
    def get(self, request):
        ctx = self.admin.each_context(request)
        
        # Get date parameters from request
        date_from_str = request.GET.get('date_from', '')
        date_to_str = request.GET.get('date_to', '')
        page_number = request.GET.get('page', 1)
        
        # Set default dates if not provided
        if not date_from_str:
            date_from = datetime.today() - timedelta(days=30)  # Default to last 30 days
        else:
            date_from = datetime.strptime(date_from_str, '%Y-%m-%d')
            
        if not date_to_str:
            date_to = datetime.today()
        else:
            date_to = datetime.strptime(date_to_str, '%Y-%m-%d')
            # Make the end date inclusive by setting it to end of day
            date_to = date_to.replace(hour=23, minute=59, second=59)
        
        # Filter orders by date range
        orders_query = Order.objects.filter(
            created_at__gte=date_from,
            created_at__lte=date_to
        ).order_by('-created_at')
        
        # Paginate the results
        paginator = Paginator(orders_query, 25)  # Show 25 orders per page
        orders_page = paginator.get_page(page_number)
        
        # Add context variables
        ctx.update({
            'orders': orders_page,
            'date_from': date_from,
            'date_to': date_to,
            'is_filtered': bool(date_from_str or date_to_str)
        })
        
        return render(request, "range-summary.html", ctx)


class ShippingProviderAnalyze(views.generic.ListView):
    admin = {}
    
    def get(self, request):
        ctx = self.admin.each_context(request)
        
        # Get date parameters from request
        date_from_str = request.GET.get('date_from', '')
        date_to_str = request.GET.get('date_to', '')
        
        # Set default dates if not provided
        if not date_from_str:
            date_from = datetime.today() - timedelta(days=30)  # Default to last 30 days
        else:
            date_from = datetime.strptime(date_from_str, '%Y-%m-%d')
            
        if not date_to_str:
            date_to = datetime.today()
        else:
            date_to = datetime.strptime(date_to_str, '%Y-%m-%d')
            # Make the end date inclusive by setting it to end of day
            date_to = date_to.replace(hour=23, minute=59, second=59)
        
        # Get shipping provider stats with sum of item weights
        provider_stats = []
        
        shipping_providers = ShippingProvider.objects.all()
        for provider in shipping_providers:
            total_weight = OrderBasket.objects.filter(
                shipping_provider=provider,
                created_at__gte=date_from,
                created_at__lte=date_to
            ).aggregate(total_weight=Sum('items_weight'))['total_weight'] or 0
            
            order_count = OrderBasket.objects.filter(
                shipping_provider=provider,
                created_at__gte=date_from,
                created_at__lte=date_to
            ).count()
            
            provider_stats.append({
                'provider_name': provider.name,
                'total_weight': total_weight,
                'order_count': order_count,
                'provider_id': provider.id
            })
        
        # Add context variables
        ctx.update({
            'provider_stats': provider_stats,
            'date_from': date_from,
            'date_to': date_to,
            'is_filtered': bool(date_from_str or date_to_str)
        })
        
        return render(request, "shipping-provider-analyze.html", ctx)
    

def export_range_summary(request):
    # Get date parameters from request
    date_from_str = request.GET.get('date_from', '')
    date_to_str = request.GET.get('date_to', '')
    
    # Set default dates if not provided
    if not date_from_str:
        date_from = datetime.today() - timedelta(days=30)
    else:
        date_from = datetime.strptime(date_from_str, '%Y-%m-%d')
        
    if not date_to_str:
        date_to = datetime.today()
    else:
        date_to = datetime.strptime(date_to_str, '%Y-%m-d')
        date_to = date_to.replace(hour=23, minute=59, second=59)
    
    # Filter orders by date range - no pagination for export
    orders = Order.objects.filter(
        created_at__gte=date_from,
        created_at__lte=date_to
    ).order_by('-created_at')
    
    # Create a workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders Report"
    
    # Add header with date range
    ws.merge_cells('A1:F1')
    header_cell = ws['A1']
    header_cell.value = f"Orders Report ({date_from.strftime('%Y-%m-%d')} to {date_to.strftime('%Y-%m-%d')})"
    header_cell.font = Font(size=14, bold=True)
    header_cell.alignment = Alignment(horizontal='center')
    
    # Create column headers
    headers = ['ID', 'Customer', 'Status', 'Total Price ($)', 'Created At', 'Notes']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}3'] = header
        ws[f'{col_letter}3'].font = Font(bold=True)
        ws[f'{col_letter}3'].fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Add data rows
    row_num = 4
    for order in orders:
        ws[f'A{row_num}'] = order.id
        ws[f'B{row_num}'] = order.customer.full_name if order.customer else 'N/A'
        ws[f'C{row_num}'] = order.get_status_display()
        ws[f'D{row_num}'] = order.total_price
        ws[f'E{row_num}'] = order.created_at.strftime('%Y-%m-%d %H:%M')
        ws[f'F{row_num}'] = order.notes if hasattr(order, 'notes') else ''
        row_num += 1
    
    # Add summary row
    ws[f'A{row_num+1}'] = f"Total Orders: {orders.count()}"
    ws[f'A{row_num+1}'].font = Font(bold=True)
    
    ws[f'C{row_num+1}'] = "Total Value:"
    ws[f'C{row_num+1}'].font = Font(bold=True)
    
    ws[f'D{row_num+1}'] = sum(order.total_price for order in orders)
    ws[f'D{row_num+1}'].font = Font(bold=True)
    
    # Auto-size columns
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"orders_report_{date_from.strftime('%Y%m%d')}_to_{date_to.strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Save the workbook to the response
    wb.save(response)
    return response


def export_shipping_provider_analyze(request):
    # Get date parameters from request
    date_from_str = request.GET.get('date_from', '')
    date_to_str = request.GET.get('date_to', '')
    
    # Set default dates if not provided
    if not date_from_str:
        date_from = datetime.today() - timedelta(days=30)
    else:
        date_from = datetime.strptime(date_from_str, '%Y-%m-%d')
        
    if not date_to_str:
        date_to = datetime.today()
    else:
        date_to = datetime.strptime(date_to_str, '%Y-%m-%d')
        date_to = date_to.replace(hour=23, minute=59, second=59)
    
    # Get shipping provider stats
    shipping_providers = ShippingProvider.objects.all()
    provider_stats = []
    
    for provider in shipping_providers:
        total_weight = OrderBasket.objects.filter(
            shipping_provider=provider,
            created_at__gte=date_from,
            created_at__lte=date_to
        ).aggregate(total_weight=Sum('items_weight'))['total_weight'] or 0
        
        order_count = OrderBasket.objects.filter(
            shipping_provider=provider,
            created_at__gte=date_from,
            created_at__lte=date_to
        ).count()
        
        provider_stats.append({
            'provider_name': provider.name,
            'total_weight': total_weight,
            'order_count': order_count
        })
    
    # Create a workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Shipping Provider Report"
    
    # Add header with date range
    ws.merge_cells('A1:C1')
    header_cell = ws['A1']
    header_cell.value = f"Shipping Provider Analysis ({date_from.strftime('%Y-%m-%d')} to {date_to.strftime('%Y-%m-%d')})"
    header_cell.font = Font(size=14, bold=True)
    header_cell.alignment = Alignment(horizontal='center')
    
    # Create column headers
    headers = ['Provider Name', 'Total Weight (kg)', 'Order Count']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}3'] = header
        ws[f'{col_letter}3'].font = Font(bold=True)
        ws[f'{col_letter}3'].fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Add data rows
    row_num = 4
    for stat in provider_stats:
        ws[f'A{row_num}'] = stat['provider_name']
        ws[f'B{row_num}'] = stat['total_weight']
        ws[f'C{row_num}'] = stat['order_count']
        row_num += 1
    
    # Add summary row
    ws[f'A{row_num+1}'] = f"Total Providers: {len(provider_stats)}"
    ws[f'A{row_num+1}'].font = Font(bold=True)
    
    ws[f'B{row_num+1}'] = sum(stat['total_weight'] for stat in provider_stats)
    ws[f'B{row_num+1}'].font = Font(bold=True)
    
    ws[f'C{row_num+1}'] = sum(stat['order_count'] for stat in provider_stats)
    ws[f'C{row_num+1}'].font = Font(bold=True)
    
    # Auto-size columns
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"shipping_provider_report_{date_from.strftime('%Y%m%d')}_to_{date_to.strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Save the workbook to the response
    wb.save(response)
    return response